import openpyxl
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import random
import string
import os
import json
# 定义函数以安全地转换值为整数，不可转换时返回0
def safe_int(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0

# 生成随机文件名
def random_filename(prefix, suffix):
    random_str = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
    return f"{prefix}_{random_str}{suffix}"

# 辅助函数用于复制样式
def copy_style(src, tgt):
    tgt.font = Font(name=src.font.name,
                    size=src.font.size,
                    bold=src.font.bold,
                    italic=src.font.italic,
                    vertAlign=src.font.vertAlign,
                    underline=src.font.underline,
                    strike=src.font.strike,
                    color=src.font.color)
    tgt.fill = PatternFill(fill_type=src.fill.fill_type,
                           start_color=src.fill.start_color,
                           end_color=src.fill.end_color)
    tgt.border = Border(left=Side(border_style=src.border.left.border_style, color=src.border.left.color),
                        right=Side(border_style=src.border.right.border_style, color=src.border.right.color),
                        top=Side(border_style=src.border.top.border_style, color=src.border.top.color),
                        bottom=Side(border_style=src.border.bottom.border_style, color=src.border.bottom.color))
    tgt.alignment = Alignment(horizontal=src.alignment.horizontal,
                              vertical=src.alignment.vertical,
                              text_rotation=src.alignment.text_rotation,
                              wrap_text=src.alignment.wrap_text,
                              shrink_to_fit=src.alignment.shrink_to_fit,
                              indent=src.alignment.indent)

# 第一个脚本的主要逻辑，处理和保存工作簿
def process_and_sort_workbook(input_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # 插入新列并设置标题
    titles = ['Building', 'Unit', 'Floor', 'Room']
    ws.insert_cols(2, len(titles))
    for i, title in enumerate(titles, start=2):
        ws.cell(row=1, column=i).value = title

    # 拆分房间信息到新列
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):
        cell = row[0]
        room = cell.value
        if room and isinstance(room, str) and room.count('-') == 2:
            building, unit, floor_room = room.split('-')
            floor, room_number = floor_room[:-2], floor_room[-2:]
            ws.cell(row=cell.row, column=2).value = building
            ws.cell(row=cell.row, column=3).value = unit
            ws.cell(row=cell.row, column=4).value = floor
            ws.cell(row=cell.row, column=5).value = room_number

    # 收集数据和样式
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        data.append([(cell.value, cell) for cell in row])

    # 对数据排序
    data.sort(key=lambda r: (safe_int(r[1][0]), safe_int(r[2][0]), safe_int(r[3][0]), safe_int(r[4][0])))

    # 删除旧数据
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 写回数据和样式
    for row_index, row in enumerate(data, start=2):
        for col_index, (value, src_cell) in enumerate(row, start=1):
            tgt_cell = ws.cell(row=row_index, column=col_index, value=value)
            copy_style(src_cell, tgt_cell)

    # 删除不需要的列
    ws.delete_cols(2, len(titles))

    # 保存排序后的工作簿至随机文件名，并返回该文件路径
    random_file_path = random_filename('SortedRoomInfo', '.xlsx')
    wb.save(random_file_path)
    print(f'Excel file sorted and saved as "{random_file_path}"')
    return random_file_path

# 第二个脚本的主要逻辑，更新数据
def update_data(input_file_name, lookup_file_name):
    try:
        input_wb = openpyxl.load_workbook(input_file_name)
        lookup_wb = openpyxl.load_workbook(lookup_file_name)
    except FileNotFoundError as e:
        print(f"无法找到文件: {e}")
        exit()

    input_ws = input_wb['房间']
    lookup_ws = lookup_wb['Sheet1']

    # 获取列名和列号的映射
    input_headers = {cell.value: get_column_letter(cell.column) for cell in input_ws[1] if cell.value is not None}
    lookup_headers = {cell.value: get_column_letter(cell.column) for cell in lookup_ws[1] if cell.value is not None}

    # 获取共有的列名
    common_headers = set(input_headers).intersection(lookup_headers)

    # 构建查找表的映射（假设“房间”和“客户”列的组合数据是唯一的）
    lookup_map = {}
    for row in lookup_ws.iter_rows(min_row=2, values_only=True):
        # 假设“房间”是第一列，"客户"是第二列
        room_customer_pair = tuple(row[:2])  # 取前两个值作为键
        if all(room_customer_pair):  # 确保键中没有空值
            lookup_map[room_customer_pair] = row

    # 遍历录入表，填充数据
    for row in input_ws.iter_rows(min_row=2):
        room_customer_pair = tuple(cell.value for cell in row[:2])
        if room_customer_pair in lookup_map:
            # 获取查找表中的完整行数据
            lookup_row_data = lookup_map[room_customer_pair]
            for header in common_headers:
                input_col_letter = input_headers[header]
                lookup_col_letter = lookup_headers[header]
                input_col_idx = column_index_from_string(input_col_letter)
                input_cell = row[input_col_idx - 1]  # 使用列索引定位单元格，列索引需要减1因为OpenPyxl索引是从1开始的
                lookup_col_idx = column_index_from_string(lookup_col_letter) - 1
                lookup_value = lookup_row_data[lookup_col_idx]
                # 如果录入表中单元格为空并且查找表中相应列有数据
                if input_cell.value is None and lookup_value:
                    input_cell.value = lookup_value


    # 保存修改后的录入表
    input_wb.save(input_file_name)
    print(f"更新后的文件已保存至: {input_file_name}")





# 读取配置文件
def load_config(config_file):
    with open(config_file, 'r') as file:
        config = json.load(file)
    return config


# 主程序，处理工作簿并更新数据，然后保存至新的文件中
def main(config_file):
    # 加载配置
    config = load_config(config_file)

    print("开始处理 Excel 文件...")
    intermediate_file_path = process_and_sort_workbook(config['source_file'])
    print("Excel 文件处理完毕。开始更新数据...")

    update_data(intermediate_file_path, config['template_file'])
    print("数据更新完毕。")

    # 重命名中间文件为最终的文件名
    final_file_path = random_filename('ProcessedRoomInfo', '.xlsx')
    os.rename(intermediate_file_path, final_file_path)
    print(f"最终文件已保存至: {final_file_path}")

# 执行主程序
if __name__ == '__main__':
    main('config.json')